import json
import re
import openpyxl


class BOQMapper:

    def __init__(
        self,
        mapping_file,
        template_name,
        network_type="perumahan",
        template_path=None
    ):

        self.network_type = network_type.lower().strip()
        self.valid_designators = set()
        self.designator_descriptions = {}

        if template_path:
            try:
                wb = openpyxl.load_workbook(template_path, data_only=True)
                ws = None
                for name in wb.sheetnames:
                    if name.upper().strip() == "BOQ":
                        ws = wb[name]
                        break
                if not ws:
                    ws = wb.active

                # Find Designator and Uraian Pekerjaan columns
                designator_col = None
                uraian_col = None
                for row in ws.iter_rows(max_row=20):
                    for cell in row:
                        if cell.value:
                            val_str = str(cell.value).strip().upper()
                            if val_str == "DESIGNATOR":
                                designator_col = cell.column
                            elif "URAIAN" in val_str:
                                uraian_col = cell.column
                    if designator_col and uraian_col:
                        break

                if designator_col:
                    for r in range(1, ws.max_row + 1):
                        val = ws.cell(row=r, column=designator_col).value
                        if val:
                            des_code = str(val).strip()
                            self.valid_designators.add(des_code)
                            
                            if uraian_col:
                                uraian_val = ws.cell(row=r, column=uraian_col).value
                                if uraian_val:
                                    # Strip prefix to map the base spec
                                    base_code = des_code
                                    if base_code.startswith("M-") or base_code.startswith("J-"):
                                        base_code = base_code[2:]
                                    self.designator_descriptions[base_code] = str(uraian_val).strip()
            except Exception as e:
                print(f"[WARN] Error loading template designators: {e}")

        with open(
            mapping_file,
            "r",
            encoding="utf-8"
        ) as f:
            all_mapping = json.load(f)

        self.mapping = all_mapping.get(
            template_name,
            {}
        )

    # =====================================
    # SPECIFICATION CLEANSING
    # =====================================

    def clean_spec(self, spec, object_type):
        if not spec:
            return ""
        spec = spec.strip().upper()
        
        # Strip M- or J- prefix if present
        if spec.startswith("M-") or spec.startswith("J-"):
            spec = spec[2:]

        if object_type == "cable":
            # Normalisasi cable specification
            # Convert AC-OF-SM-ADSS-24-D to AC-OF-SM-ADSS-24D
            spec = re.sub(r'-(\d+)-D$', r'-\1D', spec)
            spec = spec.replace(" ", "")
            return spec

        elif object_type == "odp":
            spec = spec.replace(" ", "")
            # ODP-PB-08-SOLID -> ODP-SOLID-PB-8
            match = re.match(r'ODP-PB-(\d+)-SOLID', spec)
            if match:
                cap = int(match.group(1))
                return f"ODP-SOLID-PB-{cap}"

            match_pl = re.match(r'ODP-PL-(\d+)-SOLID', spec)
            if match_pl:
                cap = int(match_pl.group(1))
                return f"ODP-SOLID-PL-{cap}"

            # ODP-PB-8 -> ODP-PB-8 (jika ada SOLID)
            match2 = re.match(r'ODP-PB-(\d+)', spec)
            if match2:
                cap = int(match2.group(1))
                if "SOLID" in spec:
                    return f"ODP-SOLID-PB-{cap}"
                return f"ODP-PB-{cap}"

            return spec

        return spec

    # =====================================
    # DESIGNATOR GENERATOR (LEGACY COMPATIBILITY)
    # =====================================

    def generate_designator(
        self,
        pattern,
        core=None
    ):
        if "{core}" in pattern:
            if core is None:
                return None
            return pattern.format(core=core)
        return pattern

    # =====================================
    # VOLUME CALCULATION (LEGACY COMPATIBILITY)
    # =====================================

    def calculate_volume(
        self,
        item,
        volume_rule
    ):
        if volume_rule == "length":
            return float(item.get("length", 0))
        if volume_rule == "count":
            return 1
        return 0

    # =====================================
    # OBJECT KEY (LEGACY COMPATIBILITY)
    # =====================================

    def get_mapping_key(
        self,
        item
    ):
        object_type = (
            item.get("object_type", "")
            .lower()
            .strip()
        )
        return object_type

    # =====================================
    # SMART MATCHING BY DESCRIPTION / URAIAN PEKERJAAN
    # =====================================

    def match_by_description(self, text, object_type):
        if not text:
            return None
        text_lower = text.lower().strip()

        # 1. Specialized keyword matching (high accuracy)
        if object_type in ['tb', 'te', 'pole', 'point']:
            is_beton = 'beton' in text_lower or 'concrete' in text_lower
            is_besi = 'besi' in text_lower or 'steel' in text_lower or 'galvanis' in text_lower or 'galvanized' in text_lower
            is_7 = '7' in text_lower or '7m' in text_lower or '7 meter' in text_lower
            is_9 = '9' in text_lower or '9m' in text_lower or '9 meter' in text_lower
            is_6 = '6' in text_lower or '6m' in text_lower or '6 meter' in text_lower
            is_kayu = 'kayu' in text_lower or 'wood' in text_lower or 'ulin' in text_lower
            
            if is_beton and is_7:
                return "PU-C7.0-150"
            elif is_beton and is_9:
                return "PU-C9.0-150"
            elif is_besi and is_7:
                return "PU-S7.0-140"
            elif is_besi and is_9:
                return "PU-S9.0-140"
            elif is_besi and is_6:
                return "PU-G6-2,5"
            elif is_kayu:
                return "PU-W7"

        if object_type == 'cable':
            is_12 = '12' in text_lower or '12c' in text_lower or '12 core' in text_lower
            is_24 = '24' in text_lower or '24c' in text_lower or '24 core' in text_lower
            is_48 = '48' in text_lower or '48c' in text_lower or '48 core' in text_lower
            is_96 = '96' in text_lower or '96c' in text_lower or '96 core' in text_lower
            if is_12:
                return "AC-OF-SM-ADSS-12D"
            elif is_24:
                return "AC-OF-SM-ADSS-24D"
            elif is_48:
                return "AC-OF-SM-ADSS-48D"
            elif is_96:
                return "AC-OF-SM-ADSS-96D"

        if object_type == 'odp':
            is_8 = '8' in text_lower or '08' in text_lower or 'solid-8' in text_lower
            is_16 = '16' in text_lower or 'solid-16' in text_lower
            if is_8:
                return "ODP-SOLID-PB-8"
            elif is_16:
                return "ODP-SOLID-PB-16"

        # 1.5 Prefix / Code similarity match (e.g. folder name is "DD-BM-HDPE", matches "DD-BM-HDPE-40-1")
        if self.valid_designators:
            text_clean = re.sub(r'[^A-Z0-9]', '', text_lower.upper())
            if text_clean and len(text_clean) >= 4:
                for d in sorted(self.valid_designators):
                    d_base = d[2:] if d.startswith("M-") or d.startswith("J-") else d
                    d_base_clean = re.sub(r'[^A-Z0-9]', '', d_base.upper())
                    if d_base_clean.startswith(text_clean):
                        return d_base

        # 2. General fuzzy search using word overlap with Uraian Pekerjaan
        words = set(re.findall(r'\b\w+\b', text_lower))
        stop_words = {'dan', 'dengan', 'yang', 'atau', 'pada', 'untuk', 'ke', 'dari', 'di', 's/d', 'sd', 'tiang', 'kabel'}
        query_words = words - stop_words
        if not query_words:
            return None

        best_match = None
        max_overlap = 0

        for base_code, desc in self.designator_descriptions.items():
            # Check prefix compatibility
            if object_type in ['tb', 'te', 'pole', 'point'] and not base_code.startswith("PU-"):
                continue
            if object_type == 'cable' and not base_code.startswith("AC-"):
                continue
            if object_type == 'odp' and not base_code.startswith("ODP-"):
                continue
            if object_type == 'odc' and not base_code.startswith("ODC-"):
                continue
            if object_type == 'closure' and not base_code.startswith("SC-"):
                continue

            desc_words = set(re.findall(r'\b\w+\b', desc.lower())) - stop_words
            overlap = len(query_words & desc_words)
            if overlap > max_overlap:
                max_overlap = overlap
                best_match = base_code

        if max_overlap >= 2:
            return best_match

        return None

    # =====================================
    # MAP OBJECTS (MAIN LOGIC)
    # =====================================

    def map_objects(
        self,
        raw_objects
    ):
        result = {}

        # =====================================
        # 1. PERHITUNGAN AKSESORIS TIANG KHUSUS & HANDHOLE
        # =====================================
        # Hitung tiang secara dinamis berdasarkan nama folder atau tipe objek
        total_tb = 0
        total_te = 0
        for item in raw_objects:
            folder_upper = str(item.get("folder", "")).upper().strip()
            is_tb_folder = folder_upper in ["TB", "TN"] or any(x in folder_upper for x in ["BARU", "NEW", "TIANG BARU"])
            is_te_folder = folder_upper in ["TE", "TL"] or any(x in folder_upper for x in ["EKSISTING", "EXISTING", "LAMA"])
            object_type = str(item.get("object_type", "")).lower().strip()

            if object_type == "tb" or is_tb_folder:
                total_tb += 1
            elif object_type == "te" or is_te_folder:
                total_te += 1
        total_poles = total_tb + total_te

        if total_poles > 0:
            # PU-AS-HL (Aksesoris Tiang Helical Performed Grip) = Total Tiang * 3
            result["M-PU-AS-HL"] = total_poles * 3
            result["J-PU-AS-HL"] = total_poles * 3

            # PU-AS-SC (Aksesoris Tiang Suspension Clamp) = Total Tiang * 2
            result["M-PU-AS-SC"] = total_poles * 2
            result["J-PU-AS-SC"] = total_poles * 2

            # STAINLESS BELT (Aksesoris Tiang Stainless Belt) = Total Tiang * 3
            result["M-STAINLESS BELT"] = total_poles * 3
            result["J-STAINLESS BELT"] = total_poles * 3

        # Hitung Handholes (HH PIT) secara dinamis (hanya jika tipe jaringan bukan 'none')
        if self.network_type != "none":
            hh_ha_count = sum(1 for item in raw_objects if any(x in str(item.get("folder", "")).upper() for x in ["HH-PIT-HA", "HANDHOLE-HA", "HH_PIT_HA"]))
            hh_odp_count = sum(1 for item in raw_objects if any(x in str(item.get("folder", "")).upper() for x in ["HH-PIT-ODP", "HANDHOLE-ODP", "HH_PIT_ODP"]))

            mult_first = 2 if self.network_type == "perumahan" else 4

            if hh_ha_count > 0:
                vol_ha = (hh_ha_count * mult_first) * 5
                result["M-HH-PIT-HA"] = vol_ha
                result["J-HH-PIT-HA"] = vol_ha

            if hh_odp_count > 0:
                vol_odp = (hh_odp_count * mult_first) * 5
                result["M-HH-PIT-P-ODP"] = vol_odp
                result["J-HH-PIT-P-ODP"] = vol_odp

        # =====================================
        # 2. PEMETAAN OBJEK UMUM DARI KML (CABLE, ODP, ODC, CLOSURE)
        # =====================================
        for item in raw_objects:
            folder = str(item.get("folder", "")).upper().strip()
            # Lewati folder yang sudah diproses di aturan khusus (atau jika tipe jaringan 'none' abaikan handholes)
            if any(x in folder for x in ["HH-PIT-HA", "HANDHOLE-HA", "HH_PIT_HA", "HH-PIT-ODP", "HANDHOLE-ODP", "HH_PIT_ODP"]):
                continue

            object_type = str(item.get("object_type", "")).lower().strip()
            name = str(item.get("name", "")).strip()
            spec_raw = str(item.get("specification", "")).strip()
            core = item.get("core")
            length = float(item.get("length", 0) or 0)

            # Coba deteksi spesifikasi dari name, folder, atau description jika spec_raw kosong
            if not spec_raw:
                name_clean = self.clean_spec(name, object_type)
                if name_clean and self.valid_designators and (f"M-{name_clean}" in self.valid_designators or f"J-{name_clean}" in self.valid_designators):
                    spec_raw = name_clean
                else:
                    folder_clean = self.clean_spec(folder, object_type)
                    if folder_clean and self.valid_designators and (f"M-{folder_clean}" in self.valid_designators or f"J-{folder_clean}" in self.valid_designators):
                        spec_raw = folder_clean
                    else:
                        # Coba cocokkan dengan Uraian Pekerjaan menggunakan description, name, atau folder
                        desc_raw = str(item.get("description", "")).strip()
                        matched_spec = self.match_by_description(desc_raw, object_type) or \
                                       self.match_by_description(name, object_type) or \
                                       self.match_by_description(folder, object_type)
                        if matched_spec:
                            spec_raw = matched_spec

            # --- CABLE ---
            if object_type == "cable":
                spec = self.clean_spec(spec_raw, "cable")
                if not spec:
                    # Fallback jika spec di deskripsi kosong
                    core_val = core if core else 24
                    spec = f"AC-OF-SM-ADSS-{core_val}D"

                m_designator = f"M-{spec}"
                j_designator = f"J-{spec}"

                for d in [m_designator, j_designator]:
                    if not self.valid_designators or d in self.valid_designators:
                        result[d] = result.get(d, 0) + length

            # --- ODP ---
            elif object_type == "odp":
                spec = self.clean_spec(spec_raw, "odp")
                if not spec:
                    core_val = core if core else 8
                    spec = f"ODP-SOLID-PB-{core_val}"

                m_designator = f"M-{spec}"
                j_designator = f"J-{spec}"

                for d in [m_designator, j_designator]:
                    if not self.valid_designators or d in self.valid_designators:
                        result[d] = result.get(d, 0) + 1

            # --- ODC ---
            elif object_type == "odc":
                core_val = core if core else 144
                base = f"ODC-C-{core_val}"
                m_designator = f"M-{base}"
                j_designator = f"J-{base}"

                for d in [m_designator, j_designator]:
                    if not self.valid_designators or d in self.valid_designators:
                        result[d] = result.get(d, 0) + 1

            # --- CLOSURE ---
            elif object_type == "closure":
                core_val = core if core else 24
                base = f"SC-OF-SM-{core_val}"
                m_designator = f"M-{base}"
                j_designator = f"J-{base}"

                for d in [m_designator, j_designator]:
                    if not self.valid_designators or d in self.valid_designators:
                        result[d] = result.get(d, 0) + 1

            # --- POLE (TB / TN / TE / TL) ---
            elif object_type in ["tb", "te"]:
                spec = self.clean_spec(spec_raw, "pole")
                if not spec:
                    if object_type == "tb":
                        # Tipe default jika tidak ada deskripsi pada tiang baru
                        spec = "PU-C7.0-150"
                    else:
                        # Tiang eksisting tanpa spesifikasi tidak dipetakan ke pengadaan tiang baru
                        spec = None

                if spec:
                    m_designator = f"M-{spec}"
                    j_designator = f"J-{spec}"

                    for d in [m_designator, j_designator]:
                        if not self.valid_designators or d in self.valid_designators:
                            result[d] = result.get(d, 0) + 1

            else:
                # Generalized fallback for other designators (e.g. BC-TR, DC-SD, DD-BM, etc.)
                spec = self.clean_spec(spec_raw, object_type)
                if spec:
                    matched_base = None
                    if not self.valid_designators:
                        matched_base = spec
                    elif f"M-{spec}" in self.valid_designators or f"J-{spec}" in self.valid_designators:
                        matched_base = spec
                    else:
                        # Try prefix matching
                        spec_clean = re.sub(r'[^A-Z0-9]', '', spec.upper())
                        if spec_clean and len(spec_clean) >= 4:
                            for d in sorted(self.valid_designators):
                                d_base = d[2:] if d.startswith("M-") or d.startswith("J-") else d
                                d_base_clean = re.sub(r'[^A-Z0-9]', '', d_base.upper())
                                if d_base_clean.startswith(spec_clean):
                                    matched_base = d_base
                                    break
                    if matched_base:
                        # Volume is length if LineString (length > 0), else 1 (Point count)
                        volume = length if length > 0 else 1
                        for prefix in ["M-", "J-"]:
                            d = f"{prefix}{matched_base}"
                            if not self.valid_designators or d in self.valid_designators:
                                result[d] = result.get(d, 0) + volume

        return result

    # =====================================
    # PRINT RESULTS
    # =====================================

    def print_result(
        self,
        result
    ):
        print()
        print("=" * 60)
        print("BOQ RESULT")
        print("=" * 60)
        for designator, volume in sorted(result.items()):
            print(f"{designator:<40}{round(volume, 2)}")
        print()

    # =====================================
    # SUMMARY
    # =====================================

    def summary(
        self,
        result
    ):
        print()
        print(f"TOTAL DESIGNATOR : {len(result)}")
        print(f"TOTAL VOLUME     : {round(sum(result.values()), 2)}")
        print()