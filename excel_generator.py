from datetime import datetime

from openpyxl import load_workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

from openpyxl.utils import (
    get_column_letter
)

from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo
)


class ExcelGenerator:

    def __init__(
        self,
        template_path
    ):

        self.template_path = (
            template_path
        )

        self.workbook = (
            load_workbook(
                template_path
            )
        )

        boq_sheet_name = None
        for name in self.workbook.sheetnames:
            if name.upper() == "BOQ":
                boq_sheet_name = name
                break

        if boq_sheet_name:
            self.sheet = self.workbook[boq_sheet_name]
        else:
            self.sheet = self.workbook.active

    # =====================================
    # FIND HEADER COLUMN
    # =====================================

    def find_header_row_and_column(
        self,
        header_name
    ):

        for row in self.sheet.iter_rows():

            for cell in row:

                if not cell.value:
                    continue

                value = str(
                    cell.value
                ).strip().upper()

                if (
                    value ==
                    header_name.upper()
                ):

                    return cell.row, cell.column

        return None, None

    # =====================================
    # FILL BOQ
    # =====================================

    def fill_boq(
        self,
        boq_result
    ):

        header_row, designator_col = (
            self.find_header_row_and_column(
                "Designator"
            )
        )

        _, vol_col = (
            self.find_header_row_and_column(
                "Vol"
            )
        )

        if not designator_col:

            raise Exception(
                "Kolom Designator tidak ditemukan"
            )

        if not vol_col:

            raise Exception(
                "Kolom Vol tidak ditemukan"
            )

        # 1. Read all designator details from the template sheet before deletion
        designator_db = {}
        for r in range(header_row + 2, self.sheet.max_row + 1):
            des = self.sheet.cell(row=r, column=designator_col).value
            if des:
                des_str = str(des).strip()
                designator_db[des_str] = {
                    "uraian": self.sheet.cell(row=r, column=3).value,
                    "satuan": self.sheet.cell(row=r, column=4).value,
                    "harga_material": self.sheet.cell(row=r, column=5).value,
                    "harga_jasa": self.sheet.cell(row=r, column=6).value,
                    "keterangan": self.sheet.cell(row=r, column=11).value,
                    "orig_row": r
                }

        # 2. Delete all static designator rows (from row 8 onwards)
        initial_max = self.sheet.max_row
        if initial_max >= header_row + 2:
            self.sheet.delete_rows(header_row + 2, initial_max - (header_row + 1))

        # 3. Filter and sort designators that have volume > 0
        active_designators = {k: v for k, v in boq_result.items() if v > 0}
        sorted_active = sorted(
            active_designators.keys(),
            key=lambda d: designator_db.get(d, {}).get("orig_row", 999999)
        )

        # 4. Write active designator rows dynamically
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        font_regular = Font(name='Arial', size=10)

        current_row = header_row + 2
        counter = 1

        for designator in sorted_active:
            details = designator_db.get(designator, {
                "uraian": "",
                "satuan": "",
                "harga_material": 0,
                "harga_jasa": 0,
                "keterangan": "",
                "orig_row": 999999
            })

            self.sheet.cell(row=current_row, column=1).value = counter
            self.sheet.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')

            self.sheet.cell(row=current_row, column=2).value = designator
            self.sheet.cell(row=current_row, column=2).alignment = Alignment(horizontal='left')

            self.sheet.cell(row=current_row, column=3).value = details["uraian"]
            self.sheet.cell(row=current_row, column=3).alignment = Alignment(horizontal='left')

            self.sheet.cell(row=current_row, column=4).value = details["satuan"]
            self.sheet.cell(row=current_row, column=4).alignment = Alignment(horizontal='center')

            self.sheet.cell(row=current_row, column=5).value = details["harga_material"]
            self.sheet.cell(row=current_row, column=5).number_format = '#,##0'
            self.sheet.cell(row=current_row, column=5).alignment = Alignment(horizontal='right')

            self.sheet.cell(row=current_row, column=6).value = details["harga_jasa"]
            self.sheet.cell(row=current_row, column=6).number_format = '#,##0'
            self.sheet.cell(row=current_row, column=6).alignment = Alignment(horizontal='right')

            self.sheet.cell(row=current_row, column=7).value = round(active_designators[designator], 2)
            self.sheet.cell(row=current_row, column=7).number_format = '#,##0.00'
            self.sheet.cell(row=current_row, column=7).alignment = Alignment(horizontal='center')

            self.sheet.cell(row=current_row, column=8).value = f"=E{current_row}*G{current_row}"
            self.sheet.cell(row=current_row, column=8).number_format = '#,##0'
            self.sheet.cell(row=current_row, column=8).alignment = Alignment(horizontal='right')

            self.sheet.cell(row=current_row, column=9).value = f"=F{current_row}*G{current_row}"
            self.sheet.cell(row=current_row, column=9).number_format = '#,##0'
            self.sheet.cell(row=current_row, column=9).alignment = Alignment(horizontal='right')

            self.sheet.cell(row=current_row, column=10).value = f"=H{current_row}+I{current_row}"
            self.sheet.cell(row=current_row, column=10).number_format = '#,##0'
            self.sheet.cell(row=current_row, column=10).alignment = Alignment(horizontal='right')

            self.sheet.cell(row=current_row, column=11).value = details["keterangan"]
            self.sheet.cell(row=current_row, column=11).alignment = Alignment(horizontal='left')

            for col in range(1, 12):
                cell = self.sheet.cell(row=current_row, column=col)
                cell.border = thin_border
                cell.font = font_regular

            current_row += 1
            counter += 1

        # 5. Write Grand Total Row
        self.sheet.cell(row=current_row, column=3).value = "TOTAL"
        self.sheet.cell(row=current_row, column=3).font = Font(name='Arial', size=10, bold=True)
        self.sheet.cell(row=current_row, column=3).alignment = Alignment(horizontal='right')

        self.sheet.cell(row=current_row, column=8).value = f"=SUM(H8:H{current_row-1})"
        self.sheet.cell(row=current_row, column=8).font = Font(name='Arial', size=10, bold=True)
        self.sheet.cell(row=current_row, column=8).number_format = '#,##0'
        self.sheet.cell(row=current_row, column=8).alignment = Alignment(horizontal='right')

        self.sheet.cell(row=current_row, column=9).value = f"=SUM(I8:I{current_row-1})"
        self.sheet.cell(row=current_row, column=9).font = Font(name='Arial', size=10, bold=True)
        self.sheet.cell(row=current_row, column=9).number_format = '#,##0'
        self.sheet.cell(row=current_row, column=9).alignment = Alignment(horizontal='right')

        self.sheet.cell(row=current_row, column=10).value = f"=SUM(J8:J{current_row-1})"
        self.sheet.cell(row=current_row, column=10).font = Font(name='Arial', size=10, bold=True)
        self.sheet.cell(row=current_row, column=10).number_format = '#,##0'
        self.sheet.cell(row=current_row, column=10).alignment = Alignment(horizontal='right')

        total_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        for col in range(1, 12):
            cell = self.sheet.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.fill = total_fill

    # =====================================
    # UPDATE BOQ HEADER
    # =====================================

    def update_boq_header(

        self,

        source_kml,

        template_name

    ):

        project_name = (

            source_kml
            .replace(".kml", "")
            .replace("_", " ")
            .upper()

        )

        self.sheet["A2"] = None

        self.sheet["A3"] = (
            f"PROJECT : {project_name}"
        )

        self.sheet["A4"] = (
            "STO : TTG"
        )

    # =====================================
    # COUNT OBJECT
    # =====================================

    def count_object_type(
        self,
        raw_objects,
        object_type
    ):

        return sum(

            1

            for item in raw_objects

            if (
                item.get(
                    "object_type"
                ) == object_type
            )

        )
    
    # =====================================
    # HEADER STYLE
    # =====================================

    def format_header(
        self,
        ws,
        row_num
    ):

        fill = PatternFill(
            "solid",
            fgColor="D9EAD3"
        )

        font = Font(
            bold=True
        )

        for cell in ws[row_num]:

            cell.fill = fill
            cell.font = font

    # =====================================
    # BORDER
    # =====================================

    def apply_border(
        self,
        ws
    ):

        border = Border(

            left=Side(
                style="thin"
            ),

            right=Side(
                style="thin"
            ),

            top=Side(
                style="thin"
            ),

            bottom=Side(
                style="thin"
            )

        )

        for row in ws.iter_rows():

            for cell in row:

                if cell.value is not None:

                    cell.border = border

    # =====================================
    # DATA KML SHEET
    # =====================================

    def create_data_kml_sheet(

        self,

        raw_objects,

        cable_spans,

        source_kml

    ):

        if "DATA_KML" in self.workbook.sheetnames:

            del self.workbook[
                "DATA_KML"
            ]

        ws = self.workbook.create_sheet(
            "DATA_KML"
        )

        current_time = datetime.now().strftime(
            "%Y-%m-%d %H:%M"
        )

        # ==========================
        # METADATA
        # ==========================

        ws.merge_cells(
            "A1:H1"
        )

        ws["A1"] = (
            "FTTH BOQ GENERATOR"
        )

        ws["A1"].font = Font(
            bold=True,
            size=18
        )

        ws["A1"].fill = PatternFill(
            "solid",
            fgColor="D9EAD3"
        )

        ws["A3"] = (
            f"Generated Time : "
            f"{current_time}"
        )

        ws["A4"] = (
            f"Source KML     : "
            f"{source_kml}"
        )

        ws["A5"] = (
            "=" * 80
        )

        row = 8

        # ==========================
        # RAW KML DATA
        # ==========================

        ws.cell(
            row=row,
            column=1
        ).value = "RAW KML DATA"

        ws.cell(
            row=row,
            column=1
        ).font = Font(
            bold=True,
            size=14
        )

        row += 2

        headers = [

            "Object Type",
            "Folder",
            "Name",
            "Core",
            "Port",
            "Length (m)",
            "From Pole",
            "To Pole"

        ]

        for c, h in enumerate(
            headers,
            start=1
        ):

            ws.cell(
                row=row,
                column=c
            ).value = h

        self.format_header(
            ws,
            row
        )

        header_row = row

        row += 1

        start_row = row

        for item in raw_objects:

            values = [

                item.get("object_type"),
                item.get("folder"),
                item.get("name"),
                item.get("core"),
                item.get("port"),
                item.get("length"),
                item.get("from_pole"),
                item.get("to_pole")

            ]

            for c, v in enumerate(
                values,
                start=1
            ):

                ws.cell(
                    row=row,
                    column=c
                ).value = v

            row += 1

        end_row = row - 1

        if end_row >= start_row:

            table = Table(

                displayName="RAWKMLTable",

                ref=f"A{header_row}:H{end_row}"

            )

            table.tableStyleInfo = (
                TableStyleInfo(
                    name="TableStyleMedium2",
                    showRowStripes=True
                )
            )

            ws.add_table(
                table
            )

        # ==========================
        # CABLE DETAIL
        # ==========================

        row += 2

        ws.cell(
            row=row,
            column=1
        ).value = "CABLE DETAIL"

        ws.cell(
            row=row,
            column=1
        ).font = Font(
            bold=True,
            size=14
        )

        row += 2

        headers = [

            "Cable",
            "Core",
            "From Pole",
            "To Pole",
            "Length (m)"

        ]

        for c, h in enumerate(
            headers,
            start=1
        ):

            ws.cell(
                row=row,
                column=c
            ).value = h

        self.format_header(
            ws,
            row
        )

        cable_header_row = row

        row += 1

        cable_start_row = row

        total_span = 0

        for span in cable_spans:

            length = float(
                span.get(
                    "length",
                    0
                ) or 0
            )

            total_span += length

            values = [

                span.get("cable"),
                span.get("core"),
                span.get("from_pole"),
                span.get("to_pole"),
                length

            ]

            for c, v in enumerate(
                values,
                start=1
            ):

                ws.cell(
                    row=row,
                    column=c
                ).value = v

            row += 1

        cable_end_row = row - 1

        if cable_end_row >= cable_start_row:

            table = Table(

                displayName="CableDetailTable",

                ref=f"A{cable_header_row}:E{cable_end_row}"

            )

            table.tableStyleInfo = (
                TableStyleInfo(
                    name="TableStyleMedium2",
                    showRowStripes=True
                )
            )

            ws.add_table(
                table
            )

        # ==========================
        # SUMMARY
        # ==========================

        row += 2

        ws.cell(
            row=row,
            column=1
        ).value = "SUMMARY"

        ws.cell(
            row=row,
            column=1
        ).font = Font(
            bold=True,
            size=14
        )

        row += 2

        ws.cell(
            row=row,
            column=1
        ).value = "Description"

        ws.cell(
            row=row,
            column=2
        ).value = "Value"

        self.format_header(
            ws,
            row
        )

        row += 1

        total_length = sum(

            float(
                item.get(
                    "length",
                    0
                ) or 0
            )

            for item in raw_objects

            if item.get(
                "object_type"
            ) == "cable"

        )

        total_tb = 0
        total_te = 0
        for item in raw_objects:
            folder_upper = str(item.get("folder", "")).upper().strip()
            is_tb_folder = folder_upper in ["TB", "TN"] or any(x in folder_upper for x in ["BARU", "NEW", "TIANG BARU"])
            is_te_folder = folder_upper in ["TE", "TL"] or any(x in folder_upper for x in ["EKSISTING", "EXISTING", "LAMA"])
            object_type = str(item.get("object_type", "")).lower().strip()

            if object_type == "tb" or is_tb_folder:
                total_tb += 1
            elif object_type == "te" or is_te_folder:
                total_te += 1

        total_pole = (
            total_tb +
            total_te
        )

        total_cable = self.count_object_type(
            raw_objects,
            "cable"
        )

        total_odp = self.count_object_type(
            raw_objects,
            "odp"
        )

        total_closure = self.count_object_type(
            raw_objects,
            "closure"
        )

        total_odc = self.count_object_type(
            raw_objects,
            "odc"
        )

        total_handhole = self.count_object_type(
            raw_objects,
            "handhole"
        )

        summary = [

            (
                "Total Pole TB",
                total_tb
            ),

            (
                "Total Pole TE",
                total_te
            ),

            (
                "Total Pole All",
                total_pole
            ),

            (
                "Total Cable",
                total_cable
            ),

            (
                "Total Cable Length (KML)",
                round(
                    total_length,
                    2
                )
            ),

            (
                "Total Cable Length (Span)",
                round(
                    total_span,
                    2
                )
            ),

            (
                "Total ODP",
                total_odp
            ),

            (
                "Total Closure",
                total_closure
            ),

            (
                "Total ODC",
                total_odc
            ),

            (
                "Total Handhole",
                total_handhole
            )

        ]

        for key, value in summary:

            ws.cell(
                row=row,
                column=1
            ).value = key

            ws.cell(
                row=row,
                column=2
            ).value = value

            row += 1

        self.apply_border(
            ws
        )

        self.auto_fit(
            ws
        )

        ws.freeze_panes = "A8"

    # =====================================
    # AUTO FIT
    # =====================================

    def auto_fit(
        self,
        worksheet
    ):

        for column in worksheet.columns:

            max_length = 0

            column_letter = (
                get_column_letter(
                    column[0].column
                )
            )

            for cell in column:

                try:

                    if cell.value:

                        max_length = max(

                            max_length,

                            len(
                                str(
                                    cell.value
                                )
                            )
                        )

                except:

                    pass

            worksheet.column_dimensions[
                column_letter
            ].width = max_length + 5

    # =====================================
    # SAVE
    # =====================================

    def save(
        self,
        output_file
    ):

        for sheet_name in list(self.workbook.sheetnames):
            name_upper = sheet_name.upper().strip()
            if name_upper != "BOQ" and name_upper != "DATA_KML":
                try:
                    del self.workbook[sheet_name]
                except:
                    pass

        try:

            self.workbook.calculation.fullCalcOnLoad = True

            self.workbook.calculation.forceFullCalc = True

        except:

            pass

        self.workbook.save(
            output_file
        )